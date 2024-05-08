from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from django.http import HttpResponse, JsonResponse
from django.db.models import QuerySet, Model
from django.db.models.options import Options
from django.utils.datastructures import ImmutableList


# def export_model(wb: Workbook, objs: QuerySet, exclude: list[str] = None) -> None:
#     meta: Options = objs[0]._meta
#     ws: Worksheet = wb.create_sheet(title=meta.verbose_name)
#     excel_title = [field.verbose_name for field in meta.fields if field.name not in exclude]
#     field_names = [field.name for field in meta.fields if field.name not in exclude]
#     ws.append(excel_title)
#     ws.append(field_names)
#     for obj in objs:
#         data = [f'{getattr(obj, field)}' for field in field_names]
#         ws.append(data)


def export_as_excel(self, request, queryset: QuerySet):
    meta = self.model._meta
    response: HttpResponse = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=Export_{datetime.now()}.xlsx'
    # field_names = [field.name for field in meta.fields]
    excel_title = [field.verbose_name for field in meta.fields]
    field_names = [field.name for field in meta.fields]
    wb = Workbook()
    ws = wb.active
    ws.append(excel_title)
    for obj in queryset:
        data = [f'{getattr(obj, field)}' for field in field_names]
        row = ws.append(data)
    wb.save(response)
    return response


def api_export_models(models: list[QuerySet[Model]], exclude: list[str] = None) -> HttpResponse:
    wb: Workbook = Workbook()
    response: HttpResponse = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment;filename=Export_{datetime.now()}.xlsx'
    for objs in models:
        meta: Options = objs[0]._meta
        ws: Worksheet = wb.create_sheet(title=meta.verbose_name)
        # excel_title = [field.verbose_name for field in meta.fields if field.name not in exclude]
        # field_names = [field.name for field in meta.fields if field.name not in exclude]
        excel_title: list[str] = []
        field_names: list[str] = []
        for field in meta.fields:
            if field.name not in exclude:
                excel_title.append(field.verbose_name)
                field_names.append(field.name)
        ws.append(excel_title)
        ws.append(field_names)
        for obj in objs:
            data = [f'{getattr(obj, field)}' for field in field_names]
            ws.append(data)
    wb.save(response)
    return response


def api_export_templates(model: Model, exclude: list[str] = None) -> HttpResponse:
    if exclude is None:
        exclude: list = []
    wb: Workbook = Workbook()
    response: HttpResponse = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment;filename=Export_templates_{datetime.now()}.xlsx'
    meta: Options = model._meta
    # ws: Worksheet = wb.create_sheet(title=meta.verbose_name)
    ws: Worksheet = wb.active
    field: ImmutableList = meta.fields
    verbose: list[str] = []
    title: list[str] = []
    for f in field:
        if f.name not in exclude:
            verbose.append(f.verbose_name)
            title.append(f.name)
    ws.append(["导入时务必将**第一行**和**第二行**删除之后再导入！如果有空值请填写 * 或者 -"])
    ws.append(verbose)
    ws.append(title)
    # for field in meta.fields:
    #     field
    # excel_title = [field.verbose_name for field in meta.fields if field.name not in exclude]
    # field_names = [field.name for field in meta.fields if field.name not in exclude]
    # ws.append(excel_title)
    # ws.append(field_names)
    # ws.append(data)
    wb.save(response)
    return response
